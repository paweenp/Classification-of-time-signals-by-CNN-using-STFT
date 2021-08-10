# read excell file

import numpy as np
import matplotlib
#import matplotlib as matplotlib
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
#os is used to work with files & directories
import os



#print(df)

for k in range(1,4):
    df = openpyxl.load_workbook("Data Object " + str(k)+".xlsx")

    sheet = df.active
    rows = sheet.max_row
    row = rows + 1
    columns = sheet.max_column
    col = columns + 1
    print(rows, columns)
    lista = []
    listb = []
    directory = "object" + str(k)
    parent_dir = os.getcwd()
    path = os.path.join(parent_dir, directory)
    os.mkdir(path)
    os.chdir(path)

    for j in range(1, row):
        for i in range(1, col):
            e = sheet.cell(row=j, column=i)
            lista.append(e.value)
            listb.append(e.value)
        folder = "Time_Domain"
        cur_dir = os.getcwd()
        cur_path = os.path.join(cur_dir, folder)
        # dir_exist = os.path.exists(curpath)
        # dir_exist = True
        if not os.path.exists(cur_path):
            os.mkdir(cur_path)
        os.chdir(cur_path)
        plt.plot(lista)
        NOP = str(j)
        plt.title(NOP)
        #plt.show()
        plt.savefig(NOP + ".jpg", bbox_inches="tight")
        plt.clf()
        lista = []
        os.chdir(path)
        folder = "Frequency_Domain"
        curdir = os.getcwd()
        curpath = os.path.join(curdir, folder)
        if not os.path.exists(curpath):
            os.mkdir(curpath)
        os.chdir(curpath)
        plt.specgram(listb)
        plt.title(NOP)
        plt.savefig(NOP + ".jpg", bbox_inches="tight")
        plt.clf()
        listb = []
        os.chdir(path)
        #plt_spec = plt.make_spectrum(lista)
        #plt_spec.plot()
        #plt_spec.show()

    os.chdir(parent_dir)






